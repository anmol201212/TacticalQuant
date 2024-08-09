from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
import time

def Get_data(url):  
    # Set options for Microsoft Edge
    options = Options()
    options.use_chromium = True  # Use this option if using Edge Chromium

    # Specify the path to the WebDriver executable
    webdriver_path = r"C:\Users\anmol.chopra\Downloads\edgedriver_win64 (1)\msedgedriver.exe"

    # Initialize the WebDriver
    service = Service(webdriver_path)
    driver = webdriver.Edge(service=service, options=options)

    # Open a particular website
    # url = 'https://www.cmegroup.com/markets/agriculture/oilseeds/soybean.volume.html'
    driver.get(url)

    time.sleep(10)
    html_source = driver.page_source

    soup = BeautifulSoup(html_source, 'html.parser')

    # Locate all tables
    tables = soup.find_all('div', {'class': 'main-table-wrapper'})

    # Check if there are enough tables
    if len(tables) > 1:
        table1 = tables[0].find('table')  # Select the first table
        table2 = tables[1].find('table')  # Select the second table
    else:
        raise ValueError("Not enough tables found on the page.")


    td_elements = table1.find_all('td')

    # Extract the text from each <td> element and remove commas
    values = [int(td.get_text().replace(',', '')) for td in td_elements]

    label_element = driver.find_element(By.XPATH, '//label[text()="Trade Date"]')

    # Find the adjacent sibling element containing the date
    date_element = label_element.find_element(By.XPATH, './following-sibling::div//span[@class="button-text"]')

    # Extract the date text from the element
    date_text = date_element.text

    # Extract table headers
    headers = []
    header_rows = table2.find_all('thead')
    for header_row in header_rows:
        cols = header_row.find_all('th')
        for col in cols:
            colspan = int(col.get('colspan', 1))
            headers.extend([col.get_text(separator=' ').strip()] * colspan)

    # Extract table rows
    rows = []
    for row in table2.find_all('tbody')[0].find_all('tr'):
        cols = row.find_all('td')
        if len(cols) > 0:
            cols = [ele.get_text().strip() for ele in cols]
            rows.append(cols)

    # Create DataFrame
    df = pd.DataFrame(rows, columns=headers[:len(rows[0])])
    driver.quit()

    return df,date_text,values[-2:]

def data_transform(df,date_text):
    df = df[['Month','Open Interest']]
    df.columns = ['Month','OI','Open Interest']
    df = df.drop(columns='OI')
    df['Open Interest'] = df['Open Interest'].str.replace(',', '').astype(int)
    df['Open Interest'] = df['Open Interest'].astype(int)
    df_new = df.T
    df_new = df_new.astype(str)
    df_new.columns = df_new.iloc[0]
    df_new = df_new.drop(df_new.index[0])
    df_new.insert(0,'Date',date_text.split(', ')[1] )
    return df_new

def cme_update():
        
    links = ['https://www.cmegroup.com/markets/agriculture/oilseeds/soybean.volume.html','https://www.cmegroup.com/markets/agriculture/oilseeds/soybean-meal.volume.html','https://www.cmegroup.com/markets/agriculture/oilseeds/soybean-oil.volume.html','https://www.cmegroup.com/markets/agriculture/grains/corn.volume.html','https://www.cmegroup.com/markets/agriculture/grains/wheat.volume.html','https://www.cmegroup.com/markets/agriculture/grains/kc-wheat.volume.html']
    products = ['Soybean','Soymeal','Soyoil','Corn','SRW','HRW']
    for i in range(0,6):
        df,date_text,values = Get_data(links[i])
        temp_df1 = data_transform(df,date_text)
        temp_df1['Market Open Interest'] = values[0]
        temp_df1['Change in OI'] = values[1]
        # temp_df1.to_csv(r'\\corp.hertshtengroup.com\Users\India\Data\anmol.chopra\Documents\code\OI\OI Data-'+products[i]+'.csv', index=False)

        temp_df2 = pd.read_csv(r'\\corp.hertshtengroup.com\Users\India\Data\anmol.chopra\Documents\FF Codes\Futures-First\OI\OI Data-'+products[i]+'.csv')
        main_df = pd.concat([temp_df1,temp_df2], ignore_index=True)
        # main_df.iloc[:, 1:] = main_df.iloc[:, 1:].astype(int)
        main_df = main_df.iloc[:,1:].apply(pd.to_numeric, errors='coerce')

        # Save the updated DataFrame to a CSV file
        main_df.to_csv(r'\\corp.hertshtengroup.com\Users\India\Data\anmol.chopra\Documents\FF Codes\Futures-First\OI\OI Data-'+products[i]+'.csv', index=False)



    df1 = pd.read_csv(r'\\corp.hertshtengroup.com\Users\India\Data\anmol.chopra\Documents\FF Codes\Futures-First\OI\OI Data-Corn.csv')
    df1.drop_duplicates(subset='Date', keep='first', inplace=True)
    df2 = pd.read_csv(r'\\corp.hertshtengroup.com\Users\India\Data\anmol.chopra\Documents\FF Codes\Futures-First\OI\OI Data-HRW.csv')
    df2.drop_duplicates(subset='Date', keep='first', inplace=True)
    df3 = pd.read_csv(r'\\corp.hertshtengroup.com\Users\India\Data\anmol.chopra\Documents\FF Codes\Futures-First\OI\OI Data-Soybean.csv')
    df3.drop_duplicates(subset='Date', keep='first', inplace=True)
    df4 = pd.read_csv(r'\\corp.hertshtengroup.com\Users\India\Data\anmol.chopra\Documents\FF Codes\Futures-First\OI\OI Data-Soymeal.csv')
    df4.drop_duplicates(subset='Date', keep='first', inplace=True)
    df5 = pd.read_csv(r'\\corp.hertshtengroup.com\Users\India\Data\anmol.chopra\Documents\FF Codes\Futures-First\OI\OI Data-Soyoil.csv')
    df5.drop_duplicates(subset='Date', keep='first', inplace=True)
    df6 = pd.read_csv(r'\\corp.hertshtengroup.com\Users\India\Data\anmol.chopra\Documents\FF Codes\Futures-First\OI\OI Data-SRW.csv')
    df6.drop_duplicates(subset='Date', keep='first', inplace=True)

    file_path = r'\\corp.hertshtengroup.com\Users\India\Data\anmol.chopra\Documents\FF Codes\Futures-First\OI\OI Data.xlsx'

    # wb = Workbook()

    # wb.save(file_path)

    # # Save dataframes to different sheets in the cleared Excel file
    # with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
    #     df1.to_excel(writer, sheet_name='Corn', index=False)
    #     df2.to_excel(writer, sheet_name='HRW', index=False)
    #     df3.to_excel(writer, sheet_name='Soybean', index=False)
    #     df4.to_excel(writer, sheet_name='Soymeal', index=False)
    #     df5.to_excel(writer, sheet_name='Soyoil', index=False)
    #     df6.to_excel(writer, sheet_name='SRW', index=False)

    # wb = load_workbook(file_path)

    # Save dataframes to different sheets in the existing Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df1.to_excel(writer, sheet_name='Corn', index=False)
        df2.to_excel(writer, sheet_name='HRW', index=False)
        df3.to_excel(writer, sheet_name='Soybean', index=False)
        df4.to_excel(writer, sheet_name='Soymeal', index=False)
        df5.to_excel(writer, sheet_name='Soyoil', index=False)
        df6.to_excel(writer, sheet_name='SRW', index=False)

def get_data_euro(url):
    
    
# Set options for Microsoft Edge
    options = Options()
    options.use_chromium = True  # Use this option if using Edge Chromium

    # Specify the path to the WebDriver executable
    webdriver_path = r"C:\Users\anmol.chopra\Downloads\edgedriver_win64 (1)\msedgedriver.exe"

    # Initialize the WebDriver
    service = Service(webdriver_path)
    driver = webdriver.Edge(service=service, options=options)

    # Open a particular website
    
    driver.get(url)

    time.sleep(15)

    html_source = driver.page_source

    soup = BeautifulSoup(html_source, 'html.parser')
    table = soup.find('table', {'id': 'future-prices-table'})

    # Extracting table headers
    headers = []
    for th in table.find_all('th'):
        headers.append(th.get_text(strip=True))

    # Extracting table rows
    rows = []
    for tr in table.find('tbody').find_all('tr'):
        cells = tr.find_all('td')
        row = [cell.get_text(strip=True) for cell in cells]
        rows.append(row)

    # Creating DataFrame
    df = pd.DataFrame(rows, columns=headers)
    date_element = soup.find('h3', class_='text-white mb-4 mb-lg-0')

# Extracting the text and splitting to get the date
    date_text = date_element.get_text(strip=True)
    date = date_text.split('Prices - ')[1]


    # Displaying the DataFrame
    # print(df)
    return df,date
    driver.quit()

def data_transform_euro(df,date,product):
    df = df[['Delivery','O.I']]
    df['O.I'] = df['O.I'].str.replace(',', '').astype(int)
    df['O.I'] = df['O.I'].astype(int)
    df_new = df.T
    df_new = df_new.astype(str)
    df_new.columns = df_new.iloc[0]
    df_new = df_new.drop(df_new.index[0])
    df_new.insert(0,'Date',date)
    df_new['Total OI'] = df['O.I'].sum()

    temp_df2 = pd.read_csv(r'\\corp.hertshtengroup.com\Users\India\Data\anmol.chopra\Documents\FF Codes\Futures-First\OI\OI Data-'+product+'.csv')
    main_df = pd.concat([df_new,temp_df2], ignore_index=True)
    main_df.iloc[:, 1:] = main_df.iloc[:, 1:].astype(int)

    # main_df.to_csv(r'\\corp.hertshtengroup.com\Users\India\Data\anmol.chopra\Documents\FF Codes\Futures-First\OI\OI Data-'+product+'.csv', index=False)
    return main_df

def transform_headers(headers):
    new_headers = []
    for header in headers:
        if header != 'Date' and header != 'Total OI':
            # Split by space and add '20' before the year
            parts = header.split()
            if len(parts) == 2:  # Format is Month Year
                month, year = parts
                new_year = '20' + year
                new_header = f"{month} {new_year}"
                new_headers.append(new_header)
        else:
            new_headers.append(header)
    return new_headers

def euronext_update():
    links = ['https://live.euronext.com/en/product/commodities-futures/EBM-DPAR','https://live.euronext.com/en/product/commodities-futures/ECO-DPAR']
    products = ['yEBM','yECO']
    file_path = r'\\corp.hertshtengroup.com\Users\India\Data\anmol.chopra\Documents\FF Codes\Futures-First\OI\OI Data.xlsx'

    for i in range(0,2):
        df,date = get_data_euro(links[i])
        df_new = data_transform_euro(df,date,products[i])
        df_new.drop_duplicates(subset='Date', keep='first', inplace=True)

        data = df_new.iloc[:,1:-1].apply(pd.to_numeric, errors='coerce')

        for j in range(len(data) - 1):
            data.iloc[j] = data.iloc[j] - data.iloc[j+1]

        df_new.iloc[:,1:-1] = data
        df_new = df_new.iloc[:-1]
        print(df_new)
        df_new.columns = transform_headers(df_new.columns)

        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_new.to_excel(writer, sheet_name=products[i], index=False)




cme_update()

euronext_update()